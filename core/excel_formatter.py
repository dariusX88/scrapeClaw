"""Professional Excel report generation with openpyxl."""

from __future__ import annotations

import logging
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rich.console import Console

from core.claude_processor import ExtractionResult
from core.config_manager import ExcelThemeConfig

logger = logging.getLogger("scrapeclaw.excel")
console = Console()


class ExcelFormatter:
    """
    Generates professional Excel reports with three worksheets:
    - Summary: run metadata, KPIs, statistics
    - Data: all extracted records with branded styling
    - Charts: category distribution bar chart + pie chart
    """

    def __init__(self, theme: ExcelThemeConfig | None = None) -> None:
        self.theme = theme or ExcelThemeConfig()

    def generate(
        self,
        results: list[ExtractionResult],
        site_name: str,
        output_path: Path,
        run_metadata: dict[str, Any] | None = None,
    ) -> Path:
        """Generate a complete Excel report."""
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        self._build_summary_sheet(wb, results, site_name, run_metadata)
        self._build_data_sheet(wb, results)
        self._build_charts_sheet(wb, results)

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        logger.info(f"Excel report saved: {output_path}")
        return output_path

    # --- Summary Sheet ---

    def _build_summary_sheet(
        self,
        wb: Workbook,
        results: list[ExtractionResult],
        site_name: str,
        run_metadata: dict[str, Any] | None = None,
    ) -> None:
        ws = wb.create_sheet("Summary")

        # Branded header banner
        ws.merge_cells("A1:E1")
        title_cell = ws["A1"]
        title_cell.value = f"ScrapeClaw Report — {site_name}"
        title_cell.font = Font(
            name="Calibri", size=18, bold=True, color=self.theme.header_font_color
        )
        title_cell.fill = PatternFill(fill_type="solid", fgColor=self.theme.header_color)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        # Subtitle row
        ws.merge_cells("A2:E2")
        ws["A2"].value = "AI-Powered Web Scraping Report"
        ws["A2"].font = Font(name="Calibri", size=11, italic=True, color="666666")
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 22

        # Section header style
        section_font = Font(name="Calibri", size=12, bold=True, color=self.theme.accent_color)
        section_fill = PatternFill(fill_type="solid", fgColor=self.theme.alt_row_color)

        # --- KPI Section ---
        ws.merge_cells("A4:B4")
        ws["A4"].value = "Key Metrics"
        ws["A4"].font = section_font
        ws["A4"].fill = section_fill
        ws["B4"].fill = section_fill

        successful = sum(1 for r in results if not r.error)
        failed = sum(1 for r in results if r.error)
        total_tokens = sum(r.tokens_used for r in results)

        kpis = [
            ("Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Site", site_name),
            ("Total Records", len(results)),
            ("Successful Extractions", successful),
            ("Failed Extractions", failed),
            ("Success Rate", f"{successful / max(len(results), 1) * 100:.1f}%"),
            ("Total Claude Tokens", f"{total_tokens:,}"),
        ]

        label_font = Font(name="Calibri", size=10, bold=True, color="333333")
        value_font = Font(name="Calibri", size=10, color="444444")
        thin_border = Border(
            bottom=Side(style="thin", color=self.theme.border_color)
        )

        for idx, (label, value) in enumerate(kpis, start=5):
            cell_label = ws.cell(row=idx, column=1, value=label)
            cell_label.font = label_font
            cell_label.border = thin_border
            cell_value = ws.cell(row=idx, column=2, value=value)
            cell_value.font = value_font
            cell_value.border = thin_border

        # Add run metadata if provided
        if run_metadata:
            meta_row = 5 + len(kpis) + 1
            ws.merge_cells(f"A{meta_row}:B{meta_row}")
            ws[f"A{meta_row}"].value = "Run Configuration"
            ws[f"A{meta_row}"].font = section_font
            ws[f"A{meta_row}"].fill = section_fill
            ws[f"B{meta_row}"].fill = section_fill

            for key, val in run_metadata.items():
                meta_row += 1
                ws.cell(row=meta_row, column=1, value=key).font = label_font
                ws.cell(row=meta_row, column=2, value=str(val)).font = value_font

        # Column widths
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 32
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18

    # --- Data Sheet ---

    def _build_data_sheet(self, wb: Workbook, results: list[ExtractionResult]) -> None:
        ws = wb.create_sheet("Data")

        if not results:
            ws["A1"] = "No data extracted."
            ws["A1"].font = Font(italic=True, color="999999")
            return

        # Collect all field names (union of all results), excluding internal fields
        display_fields = self._get_display_fields(results)
        all_columns = ["URL"] + display_fields + ["Error"]

        # Header row with branded styling
        header_fill = PatternFill(fill_type="solid", fgColor=self.theme.header_color)
        header_font = Font(name="Calibri", size=10, bold=True, color=self.theme.header_font_color)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, col_name in enumerate(all_columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        ws.row_dimensions[1].height = 28

        # Data rows with alternating colors and conditional formatting
        alt_fill = PatternFill(fill_type="solid", fgColor=self.theme.alt_row_color)
        error_fill = PatternFill(fill_type="solid", fgColor="FFEBEE")  # light red
        data_font = Font(name="Calibri", size=10, color="333333")
        url_font = Font(name="Calibri", size=10, color=self.theme.accent_color, underline="single")
        error_font = Font(name="Calibri", size=10, color="C0392B")
        thin_border = Border(
            bottom=Side(style="thin", color=self.theme.border_color)
        )

        for row_idx, result in enumerate(results, start=2):
            is_error = bool(result.error)
            is_alt = row_idx % 2 == 0

            for col_idx, col_name in enumerate(all_columns, start=1):
                if col_name == "URL":
                    value = result.url
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = url_font
                    # Make URL clickable
                    cell.hyperlink = value
                elif col_name == "Error":
                    value = result.error or ""
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = error_font if value else data_font
                else:
                    value = result.data.get(col_name, "")
                    if value is None:
                        value = ""
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = data_font

                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

                # Row coloring
                if is_error:
                    cell.fill = error_fill
                elif is_alt:
                    cell.fill = alt_fill

        # Auto-size columns (cap at 50 chars)
        for col_idx, col_name in enumerate(all_columns, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(col_name))
            for result in results[:50]:  # Sample first 50 for sizing
                if col_name == "URL":
                    val = result.url
                elif col_name == "Error":
                    val = result.error or ""
                else:
                    val = str(result.data.get(col_name, "") or "")
                max_len = max(max_len, min(len(val), 50))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

        # Freeze header row + auto-filter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    # --- Charts Sheet ---

    def _build_charts_sheet(self, wb: Workbook, results: list[ExtractionResult]) -> None:
        ws = wb.create_sheet("Charts")

        # Sheet title
        ws.merge_cells("A1:F1")
        ws["A1"].value = "Data Visualizations"
        ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=self.theme.header_color)
        ws.row_dimensions[1].height = 35

        valid_results = [r for r in results if not r.error]

        if not valid_results:
            ws["A3"] = "No data available for charts."
            ws["A3"].font = Font(italic=True, color="999999")
            return

        # --- Category Distribution (Bar Chart) ---
        categories = Counter(r.data.get("_category", "Extracted") for r in valid_results)

        ws["A3"] = "Category"
        ws["B3"] = "Count"
        ws["A3"].font = Font(bold=True)
        ws["B3"].font = Font(bold=True)

        header_fill = PatternFill(fill_type="solid", fgColor=self.theme.alt_row_color)
        ws["A3"].fill = header_fill
        ws["B3"].fill = header_fill

        for row_idx, (cat, count) in enumerate(sorted(categories.items()), start=4):
            ws.cell(row=row_idx, column=1, value=cat)
            ws.cell(row=row_idx, column=2, value=count)

        last_row = 3 + len(categories)

        if len(categories) > 1:
            # Bar chart
            bar_chart = BarChart()
            bar_chart.type = "col"
            bar_chart.title = "Records by Category"
            bar_chart.y_axis.title = "Count"
            bar_chart.x_axis.title = "Category"
            bar_chart.style = 10
            bar_chart.width = 20
            bar_chart.height = 12

            data_ref = Reference(ws, min_col=2, min_row=3, max_row=last_row)
            cats_ref = Reference(ws, min_col=1, min_row=4, max_row=last_row)
            bar_chart.add_data(data_ref, titles_from_data=True)
            bar_chart.set_categories(cats_ref)
            bar_chart.shape = 4
            ws.add_chart(bar_chart, "D3")

            # Pie chart
            pie_chart = PieChart()
            pie_chart.title = "Category Distribution"
            pie_chart.style = 10
            pie_chart.width = 16
            pie_chart.height = 12

            pie_data = Reference(ws, min_col=2, min_row=3, max_row=last_row)
            pie_cats = Reference(ws, min_col=1, min_row=4, max_row=last_row)
            pie_chart.add_data(pie_data, titles_from_data=True)
            pie_chart.set_categories(pie_cats)
            pie_chart.dataLabels = DataLabelList()
            pie_chart.dataLabels.showPercent = True
            ws.add_chart(pie_chart, "D20")

        # --- Field Completeness Stats ---
        stats_start_row = last_row + 3
        ws.cell(row=stats_start_row, column=1, value="Field Completeness")
        ws.cell(row=stats_start_row, column=1).font = Font(
            name="Calibri", size=12, bold=True, color=self.theme.accent_color
        )

        ws.cell(row=stats_start_row + 1, column=1, value="Field").font = Font(bold=True)
        ws.cell(row=stats_start_row + 1, column=2, value="Filled").font = Font(bold=True)
        ws.cell(row=stats_start_row + 1, column=3, value="Empty").font = Font(bold=True)
        ws.cell(row=stats_start_row + 1, column=4, value="Fill Rate").font = Font(bold=True)

        display_fields = self._get_display_fields(results)
        for fidx, field_name in enumerate(display_fields, start=stats_start_row + 2):
            filled = sum(
                1
                for r in valid_results
                if r.data.get(field_name) is not None and r.data.get(field_name) != ""
            )
            empty = len(valid_results) - filled
            rate = filled / max(len(valid_results), 1)

            ws.cell(row=fidx, column=1, value=field_name)
            ws.cell(row=fidx, column=2, value=filled)
            ws.cell(row=fidx, column=3, value=empty)
            rate_cell = ws.cell(row=fidx, column=4, value=f"{rate * 100:.0f}%")

            # Color code fill rate
            if rate >= 0.8:
                rate_cell.font = Font(color="27AE60", bold=True)  # green
            elif rate >= 0.5:
                rate_cell.font = Font(color="F39C12", bold=True)  # orange
            else:
                rate_cell.font = Font(color="C0392B", bold=True)  # red

        # Column widths
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 14

    # --- Helpers ---

    def _get_display_fields(self, results: list[ExtractionResult]) -> list[str]:
        """Get ordered list of non-internal field names across all results."""
        seen: list[str] = []
        for r in results:
            for key in r.data:
                if not key.startswith("_") and key not in seen:
                    seen.append(key)
        return seen
